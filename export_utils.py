import csv
import os
from datetime import datetime
from tkinter import filedialog, messagebox
import pandas as pd
from fpdf import FPDF


class ExportManager:
    """Handles exporting findings data to various formats."""
    
    def __init__(self):
        self.downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
        
    def get_timestamp(self):
        """Generate timestamp for unique filenames."""
        return datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def _flatten_finding_for_csv(self, finding):
        """Convert new schema finding to flat CSV row."""
        # Extract services into readable format
        services = finding.get('services', [])
        service_str = []
        for service in services:
            protocol = service.get('protocol', '')
            ports = service.get('ports', [])
            port_ranges = []
            for port in ports:
                port_from = port.get('from')
                port_to = port.get('to')
                if port_from == port_to:
                    port_ranges.append(str(port_from))
                else:
                    port_ranges.append(f"{port_from}-{port_to}")
            
            if port_ranges:
                service_str.append(f"{protocol}/{','.join(port_ranges)}")
            else:
                # Handle case where protocol is "any" without specific ports
                service_str.append(protocol)
        
        # Handle optional fields that might be null
        vendor = finding.get('vendor')
        name = finding.get('name')
        comments = finding.get('comments')
        evidence = finding.get('evidence', {})
        
        return {
            "Rule ID": finding.get('rule_id', ''),
            "Check ID": finding.get('check_id', ''),
            "Title": finding.get('title', ''),
            "Severity": finding.get('severity', '').capitalize(),  # Ensure proper capitalisation for {low, medium, high, critical}
            "Reason": finding.get('reason', ''),
            "Recommendation": finding.get('recommendation', ''),
            "Source Addresses": ', '.join(finding.get('src_addrs', [])),
            "Destination Addresses": ', '.join(finding.get('dst_addrs', [])),
            "Services": ', '.join(service_str),
            "Vendor": vendor.capitalize() if vendor else '',
            "Rule Name": name if name is not None else '',
            "Comments": comments if comments is not None else '',
            "Policy Name": evidence.get('policy_name', '') if evidence else '',
            "Hit Count": evidence.get('hit_count', '') if evidence else '',
            "Labels": ', '.join(finding.get('labels', []))
        }
    
    def export_to_csv(self, data, filename=None, show_dialog=True):
        """
        Export data to CSV format.
        
        Args:
            data: List of dictionaries containing the findings data (new schema format)
            filename: Optional filename, if None will use timestamp
            show_dialog: Whether to show file save dialog
        """
        try:
            if show_dialog:
                filepath = filedialog.asksaveasfilename(
                    defaultextension=".csv",
                    filetypes=[("CSV files", "*.csv")],
                    initialdir=self.downloads_folder
                )
                if not filepath:
                    return None
            else:
                if filename is None:
                    filename = f"FireFind_Results_{self.get_timestamp()}.csv"
                filepath = os.path.join(self.downloads_folder, filename)
            
            # Write CSV file with new schema
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                if data:
                    # Convert new schema to flattened CSV format
                    csv_data = []
                    for item in data:
                        csv_row = self._flatten_finding_for_csv(item)
                        csv_data.append(csv_row)
                    
                    # Use consistent headers for new schema
                    headers = ["Rule ID", "Check ID", "Title", "Severity", "Reason", "Recommendation", 
                              "Source Addresses", "Destination Addresses", "Services", "Vendor", 
                              "Rule Name", "Comments", "Policy Name", "Hit Count", "Labels"]
                    
                    writer = csv.DictWriter(f, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(csv_data)
                else:
                    # Write empty file with standard headers
                    writer = csv.writer(f)
                    writer.writerow(["Rule ID", "Check ID", "Title", "Severity", "Reason", "Recommendation", 
                                   "Source Addresses", "Destination Addresses", "Services", "Vendor", 
                                   "Rule Name", "Comments", "Policy Name", "Hit Count", "Labels"])
            
            messagebox.showinfo("Export Complete", f"CSV exported successfully to:\n{filepath}")
            return filepath
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export CSV:\n{str(e)}")
            return None
    
    def export_to_excel(self, data, filename=None, show_dialog=True):
        """
        Export data to Excel format.
        
        Args:
            data: List of dictionaries containing the findings data
            filename: Optional filename, if None will use timestamp
            show_dialog: Whether to show file save dialog
        """
        try:
            if show_dialog:
                filepath = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialdir=self.downloads_folder
                )
                if not filepath:
                    return None
            else:
                if filename is None:
                    filename = f"FireFind_Results_{self.get_timestamp()}.xlsx"
                filepath = os.path.join(self.downloads_folder, filename)
            
            # Create DataFrame and export to Excel
            if data:
                # Convert new schema to flattened format for Excel
                excel_data = [self._flatten_finding_for_csv(item) for item in data]
                df = pd.DataFrame(excel_data)
            else:
                # Create empty DataFrame with standard columns
                df = pd.DataFrame(columns=["Rule ID", "Check ID", "Title", "Severity", "Reason", "Recommendation", 
                                         "Source Addresses", "Destination Addresses", "Services", "Vendor", 
                                         "Rule Name", "Comments", "Policy Name", "Hit Count", "Labels"])
            
            # Write to Excel with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='FireFind Results', index=False)
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['FireFind Results']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format header row
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
            
            messagebox.showinfo("Export Complete", f"Excel file exported successfully to:\n{filepath}")
            return filepath
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export Excel:\n{str(e)}")
            return None
    
    def export_to_pdf(self, data, filename=None, show_dialog=True):
        """
        Export data to PDF format with professional formatting.
        
        Args:
            data: List of dictionaries containing the findings data
            filename: Optional filename, if None will use timestamp
            show_dialog: Whether to show file save dialog
        """
        try:
            if show_dialog:
                filepath = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf")],
                    initialdir=self.downloads_folder
                )
                if not filepath:
                    return None
            else:
                if filename is None:
                    filename = f"FireFind_Report_{self.get_timestamp()}.pdf"
                filepath = os.path.join(self.downloads_folder, filename)
            
            # Create PDF
            pdf = FireFindPDF()
            pdf.add_page()
            
            # Add findings
            if data:
                pdf.add_findings_table(data)
            else:
                pdf.set_font('Helvetica', '', 12)
                pdf.cell(0, 10, 'No findings to report.', new_x="LMARGIN", new_y="NEXT")
            
            # Output PDF
            pdf.output(filepath)
            
            messagebox.showinfo("Export Complete", f"PDF report exported successfully to:\n{filepath}")
            return filepath
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export PDF:\n{str(e)}")
            return None


class FireFindPDF(FPDF):
    """Custom PDF class for FireFind reports."""
    
    def __init__(self):
        super().__init__()
        self.report_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def header(self):
        """Add header to each page."""
        # Try to add FireFind logo if it exists
        logo_path = "FireFind.png"
        if os.path.exists(logo_path):
            self.image(logo_path, x=10, y=8, w=10, h=10)
        
        # Title
        self.set_font('Helvetica', 'B', 20)
        self.cell(0, 10, 'FireFind Security Report', new_x="LMARGIN", new_y="NEXT", align='C')
        
        # Subtitle and timestamp
        self.set_font('Helvetica', '', 12)
        self.cell(0, 10, f'Generated on: {self.report_datetime}', new_x="LMARGIN", new_y="NEXT", align='C')
        self.ln(10)
    
    def footer(self):
        """Add footer to each page."""
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', align='C')
    
    def add_findings_table(self, data):
        """Add findings data as a formatted table."""
        self.set_font('Helvetica', 'B', 12)
        self.cell(0, 10, 'Security Findings Summary', new_x="LMARGIN", new_y="NEXT", align='L')
        self.ln(5)
        
        # Count findings by severity
        severity_counts = {}
        for finding in data:
            severity = finding.get('severity', 'Unknown').title()
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        # Add summary
        self.set_font('Helvetica', '', 10)
        self.cell(0, 6, f'Total Findings: {len(data)}', new_x="LMARGIN", new_y="NEXT", align='L')
        for severity, count in sorted(severity_counts.items()):
            self.cell(0, 6, f'{severity}: {count}', new_x="LMARGIN", new_y="NEXT", align='L')
        self.ln(5)
        
        # Table headers for new schema
        self.set_font('Helvetica', 'B', 7)
        col_widths = [20, 45, 30, 30, 25, 20, 25]  # Column widths
        headers = ['Rule ID', 'Title', 'Source', 'Destination', 'Service', 'Severity', 'Vendor']
        
        for i, header in enumerate(headers):
            self.cell(col_widths[i], 8, header, border=1, align='C')
        self.ln()
        
        # Table data
        self.set_font('Helvetica', '', 6)
        for finding in data:
            # Handle page breaks
            if self.get_y() > 250:  # Near bottom of page
                self.add_page()
                # Repeat headers on new page
                self.set_font('Helvetica', 'B', 7)
                for i, header in enumerate(headers):
                    self.cell(col_widths[i], 8, header, border=1, align='C')
                self.ln()
                self.set_font('Helvetica', '', 6)
            
            # Extract services for display
            services = finding.get('services', [])
            service_display = []
            for service in services:
                protocol = service.get('protocol', '')
                ports = service.get('ports', [])
                port_ranges = []
                for port in ports:
                    port_from = port.get('from')
                    port_to = port.get('to')
                    if port_from == port_to:
                        port_ranges.append(str(port_from))
                    else:
                        port_ranges.append(f"{port_from}-{port_to}")
                
                if port_ranges:
                    service_display.append(f"{protocol}/{','.join(port_ranges)}")
                elif protocol:  # Handle "any" protocol case
                    service_display.append(protocol)
            
            # Handle optional fields that might be null
            vendor = finding.get('vendor')
            
            # Row data for new schema
            row_data = [
                str(finding.get('rule_id', '')),
                str(finding.get('title', ''))[:30],  # Truncate long titles
                ', '.join(finding.get('src_addrs', []))[:20],
                ', '.join(finding.get('dst_addrs', []))[:20],
                ', '.join(service_display)[:15],
                str(finding.get('severity', '')).capitalize(),
                str(vendor).capitalize() if vendor else ''
            ]
            
            for i, data_item in enumerate(row_data):
                self.cell(col_widths[i], 8, data_item, border=1, align='L')
            self.ln()


def demo_export():
    """Demo function to test export functionality."""
    sample_data = [
        {
            "id": "R-001",
            "finding": "Exposed RDP Port to Internet",
            "src": "0.0.0.0/0 (Any)",
            "dst": "10.0.2.20",
            "port": "3389/tcp",
            "rationale": "RDP exposed to internet",
            "severity": "High"
        },
        {
            "id": "R-002",
            "finding": "Permissive 'Any-Any' Rule Detected",
            "src": "0.0.0.0/0 (Any)",
            "dst": "0.0.0.0/0 (Any)",
            "port": "Any",
            "rationale": "Allow-any source/dest",
            "severity": "High"
        }
    ]
    
    exporter = ExportManager()
    print("Testing export functionality...")
    exporter.export_to_csv(sample_data, show_dialog=False)
    exporter.export_to_excel(sample_data, show_dialog=False)
    exporter.export_to_pdf(sample_data, show_dialog=False)


if __name__ == "__main__":
    demo_export()