# firefind/parsers/xlsx_parser.py
from __future__ import annotations
from openpyxl import load_workbook

REQUIRED_KEYS = {"id", "action"}  # minimal keys to confirm header row

# Accept slight header wording differences
HEADER_ALIASES = {
    "id": {"id", "policy id", "policyid"},
    "name": {"name", "policy name"},
    "action": {"action"},
    "src_addr": {
        "address | user | device", "source", "src addr", "src address", "src addresses",
        "source address", "source addresses"
    },
    "dst_addr": {
        "address", "destination", "dst addr", "dst address", "dst addresses",
        "destination address", "destination addresses"
    },
    "service": {"service | name", "service", "services"},
    "comments": {"comments", "comment", "remarks", "notes"}
}

WELL_KNOWN_SERVICES = {
    "ssh": 22,
    "rdp": 3389,
    "vnc": 5900,
    "http": 80,
    "https": 443,
}

def _norm(s) -> str:
    return "" if s is None else str(s).strip()

def _lower(s) -> str:
    return _norm(s).lower()

def _split_multi(val):
    if val is None:
        return []
    s = str(val).replace("\r", "").strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split("\n") if p.strip()]
    if not parts and "," in s:
        parts = [p.strip() for p in s.split(",") if p.strip()]
    # normalize “any”-like tokens
    if len(parts) == 1 and parts[0].lower() in {"any", "all", "0.0.0.0/0", "::/0"}:
        return ["any"]
    return parts

def _map_action(val) -> str:
    a = _lower(val)
    if a in {"accept", "allow", "permit"}: return "allow"
    if a in {"deny", "drop", "block"}:     return "deny"
    if a in {"reject"}:                    return "reject"
    return "other"

def _parse_services(cell):
    tokens = _split_multi(cell)
    if not tokens:
        return [{"protocol": "any", "ports": []}]
    out = []
    for t in tokens:
        key = t.lower()
        if key in WELL_KNOWN_SERVICES:
            p = WELL_KNOWN_SERVICES[key]
            out.append({"protocol": "tcp", "ports": [{"from": p, "to": p}]})
        elif "/" in key:  # e.g., tcp/8080
            try:
                proto, port = key.split("/", 1)
                port = int(port)
                out.append({"protocol": proto.lower(), "ports": [{"from": port, "to": port}]})
            except Exception:
                out.append({"protocol": "any", "ports": []})
        else:
            # unknown group/service name → generic for MVP
            out.append({"protocol": "any", "ports": []})
    return out

def _find_header_row(ws):
    """
    Scan first 50 rows; return (row_index, headers_list) where row_index is 0-based
    for the header row. We consider a row to be the header if, after normalizing,
    it contains the required keys (ID, Action) and at least one of service/addr columns.
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True)):
        headers = [_norm(c) for c in row]
        lower = [_lower(c) for c in headers]
        present = set()
        if any(h in HEADER_ALIASES["id"] for h in lower): present.add("id")
        if any(h in HEADER_ALIASES["action"] for h in lower): present.add("action")
        if any(h in HEADER_ALIASES["service"] for h in lower): present.add("service")
        if any(h in HEADER_ALIASES["src_addr"] for h in lower): present.add("src_addr")
        if any(h in HEADER_ALIASES["dst_addr"] for h in lower): present.add("dst_addr")

        if REQUIRED_KEYS.issubset(present) and ("service" in present or "src_addr" in present or "dst_addr" in present):
            # Trim trailing empty headers
            while headers and headers[-1] == "":
                headers.pop()
            return i, headers
    # fallback: first row
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_norm(c) for c in first]
    return 0, headers

def _build_index_map(headers):
    """Map canonical keys → column index by alias matching."""
    lower = [_lower(h) for h in headers]
    def find_idx(alias_set):
        for j, h in enumerate(lower):
            if h in alias_set:
                return j
        return None

    return {
        "id":    find_idx(HEADER_ALIASES["id"]),
        "name":  find_idx(HEADER_ALIASES["name"]),
        "action":find_idx(HEADER_ALIASES["action"]),
        "src":   find_idx(HEADER_ALIASES["src_addr"]),
        "dst":   find_idx(HEADER_ALIASES["dst_addr"]),
        "svc":   find_idx(HEADER_ALIASES["service"]),
        "cmt":   find_idx(HEADER_ALIASES["comments"]),
    }

class XlsxParser:
    def parse(self, path: str):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        header_row, headers = _find_header_row(ws)
        idx = _build_index_map(headers)

        # If we didn't even find ID & Action, nothing to do
        if idx["id"] is None or idx["action"] is None:
            return []

        rules = []
        # iterate rows AFTER header_row
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            # Skip completely empty lines
            if all(c in (None, "", " ") for c in row):
                continue

            rid = row[idx["id"]] if idx["id"] is not None and idx["id"] < len(headers) else None
            action = row[idx["action"]] if idx["action"] is not None and idx["action"] < len(headers) else None

            # Non-rule "section" rows typically have no ID
            if rid in (None, "", 0):
                continue

            rec = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}

            name = row[idx["name"]] if idx["name"] is not None and idx["name"] < len(row) else None
            src  = row[idx["src"]]  if idx["src"]  is not None and idx["src"]  < len(row) else None
            dst  = row[idx["dst"]]  if idx["dst"]  is not None and idx["dst"]  < len(row) else None
            svc  = row[idx["svc"]]  if idx["svc"]  is not None and idx["svc"]  < len(row) else None
            cmt  = row[idx["cmt"]]  if idx["cmt"]  is not None and idx["cmt"]  < len(row) else None

            rule = {
                "rule_id": str(rid),
                "vendor": "fortinet",     # for your current dataset
                "name": name,
                "enabled": True,
                "action": _map_action(action),
                "src_addrs": _split_multi(src),
                "dst_addrs": _split_multi(dst),
                "services": _parse_services(svc),
                "comments": cmt,
                "raw": rec,
            }
            rules.append(rule)

        return rules
