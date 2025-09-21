#!/usr/bin/env python3
"""
xlsx_to_csv.py : turns an .xlsx file into CSV filesss, one per sheet.
- uses openpyxl with data_only=True so we grab the last saved result of formulas
- if a sheet has merged cells, we copy the top-left value into the whole merged area
- output file names are based on the sheet name: <outdir>/<sanitized_sheet_name>.csv
"""

import argparse
import csv
import os
import re
from pathlib import Path
from openpyxl import load_workbook

def sanitize_sheet_name(name: str) -> str:
    # make something safe for a filename and not super long
    s = re.sub(r'[^\w\-.]+', '_', name.strip())
    return s[:150] or "sheet"

def build_merged_fill_map(ws):
    """
    builds a map like (row, col) -> (top_left_row, top_left_col)
    for every cell inside each merged range (including the top-left cell)
    """
    m = {}
    for mr in ws.merged_cells.ranges:
        minr, minc, maxr, maxc = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        for r in range(minr, maxr + 1):
            for c in range(minc, maxc + 1):
                m[(r, c)] = (minr, minc)
    return m

def cell_display_value(ws, r, c, merge_map):
    cell = ws.cell(row=r, column=c)
    val = cell.value
    # if this cell is part of a merged block and it's empty, grab the value from the top-left of that block
    if (r, c) in merge_map and (val is None or val == ""):
        tl_r, tl_c = merge_map[(r, c)]
        val = ws.cell(row=tl_r, column=tl_c).value
    # turn None into an empty string; everything else into a string
    if val is None:
        return ""
    return str(val)

def export_sheet(ws, out_path, delimiter, encoding, lineterminator):
    # figure out how many rows and columns the sheet uses
    max_row = ws.max_row
    max_col = ws.max_column
    merge_map = build_merged_fill_map(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding=encoding) as f:
        writer = csv.writer(
            f,
            delimiter=delimiter,
            lineterminator=lineterminator,
            quoting=csv.QUOTE_MINIMAL
        )
        for r in range(1, max_row + 1):
            row_vals = [cell_display_value(ws, r, c, merge_map) for c in range(1, max_col + 1)]
            writer.writerow(row_vals)

def main():
    ap = argparse.ArgumentParser(description="Export XLSX to CSV files (one per sheet), keeping headers and merged cells tidy.")
    ap.add_argument("xlsx", help="input .xlsx file")
    ap.add_argument("-o", "--outdir", default=".", help="where to put the CSV files (default: current folder)")
    ap.add_argument("--sheet", help="only export this exact sheet name; if not set, export all sheets")
    ap.add_argument("--delimiter", default=",", help="CSV delimiter (default: ,)")
    ap.add_argument("--encoding", default="utf-8-sig", help="output encoding (default: utf-8-sig so Excel opens it nicely)")
    ap.add_argument("--lineterminator", default="\n", help="line ending to use (default: \\n)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"Input not found: {xlsx_path}")

    # data_only=True means we read the stored results of formulas, not the formulas themselves
    wb = load_workbook(filename=str(xlsx_path), data_only=True)

    sheets = [wb[args.sheet]] if args.sheet else wb.worksheets
    for ws in sheets:
        sheet_name = ws.title
        outfile = Path(args.outdir) / f"{sanitize_sheet_name(sheet_name)}.csv"
        export_sheet(ws, outfile, args.delimiter, args.encoding, args.lineterminator)
        print(f"Wrote: {outfile}")

if __name__ == "__main__":
    main()
